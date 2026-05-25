"""Extrai schema do Excel/VBA e tabelas Access (uso local — paths via argumentos)."""
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pyodbc


def main() -> int:
    if len(sys.argv) < 3:
        print("Uso: python extract_schema.py <caminho.xlsm> <caminho.accdb>")
        return 1

    xlsm = Path(sys.argv[1])
    accdb = Path(sys.argv[2])

    if not xlsm.is_file():
        print(f"Excel não encontrado: {xlsm}")
        return 1

    print("=== EXCEL Consulta_Lancamentos (linhas 8-12) ===")
    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    ws = wb["Consulta_Lancamentos"]
    for i, row in enumerate(ws.iter_rows(min_row=8, max_row=12, values_only=True), 8):
        vals = [str(v).strip() if v is not None else "" for v in row]
        print(i, [v for v in vals if v][:25])

    print("\n=== EXCEL Consulta_DRE (primeiras linhas com texto) ===")
    ws2 = wb["Consulta_DRE"]
    for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        vals = [str(v).strip() for v in row if v not in (None, "")]
        if vals:
            print(i, vals[:15])
    wb.close()

    print("\n=== VBA strings BC_* e campos ===")
    with zipfile.ZipFile(xlsm) as z:
        vba = z.read("xl/vbaProject.bin").decode("latin-1", errors="ignore")

    for pat in [r"BC_\w+", r"TB_\w+"]:
        found = sorted(set(re.findall(pat, vba)))
        print(f"{pat}: {len(found)} ocorrências")
        for s in found[:30]:
            print(" ", s)

    if accdb.is_file():
        print("\n=== ACCESS tabelas ===")
        conn = pyodbc.connect(
            r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=" + str(accdb) + ";"
        )
        cur = conn.cursor()
        tables = [r.table_name for r in cur.tables(tableType="TABLE")]
        for t in sorted(tables):
            if not t.startswith("MSys"):
                print(t)
        conn.close()
    else:
        print(f"\nAccess não encontrado (opcional): {accdb}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
